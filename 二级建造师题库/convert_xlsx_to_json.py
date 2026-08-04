#!/usr/bin/env python3
"""Convert 二级建造师 xlsx to JSON files organized by subject and chapter."""

import re
import json
import os
from openpyxl import load_workbook
from collections import OrderedDict

TYPE_MAP = {
    '单项选择题': 'singleChoice',
    '多项选择题': 'multipleChoice',
    '判断题': 'trueFalse',
}

DIFFICULTY_MAP = {
    '简单': 'easy',
    '中等': 'medium',
    '困难': 'hard',
}

def extract_chapter(source_str):
    """Extract chapter from source. Returns (chapter_num, display_name, is_exam_group)"""
    if not source_str:
        return '0', '其他题目', False
    s = str(source_str).strip()
    m = re.match(r'第(\d+)章\s+(.+)', s)
    if m:
        return m.group(1), s, False
    return '0', s, True

def parse_answer(answer_str, qtype):
    answer_str = str(answer_str).strip().upper()
    if qtype == 'trueFalse':
        return {'isCorrect': True if answer_str in ('对', '正确', 'A', '√', '是', 'TRUE') else False}
    if qtype == 'singleChoice':
        idx = ord(answer_str[0]) - ord('A') if answer_str and answer_str[0].isalpha() else 0
        return {'answerIndex': max(0, min(idx, 3))}
    if qtype == 'multipleChoice':
        indices = [ord(ch) - ord('A') for ch in answer_str if ch.isalpha() and 0 <= (ord(ch) - ord('A')) <= 4]
        return {'answerIndices': indices}
    return {}

def clean_str(s):
    """Clean up string for JSON"""
    if s is None:
        return ''
    s = str(s).strip()
    s = re.sub(r'\s+', ' ', s)
    s = re.sub(r'<br\s*/?\s*>', ' ', s, flags=re.IGNORECASE)
    return s

def convert_xlsx(xlsx_path, subject_enum, output_dir):
    """Convert xlsx to per-chapter JSON files"""
    print(f"Reading: {xlsx_path}")
    wb = load_workbook(xlsx_path)
    ws = wb.active
    
    # Group by chapter/source
    chapters = OrderedDict()
    exam_groups = OrderedDict()
    
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[3] and row[1] and str(row[1]).strip() in TYPE_MAP:
            qtype = TYPE_MAP[str(row[1]).strip()]
            ch_num, ch_name, is_exam = extract_chapter(row[2])
            
            options = []
            for opt_col in [4, 5, 6, 7]:
                if row[opt_col]:
                    options.append(clean_str(row[opt_col]))
            if qtype == 'multipleChoice' and row[8]:
                options.append(clean_str(row[8]))
            
            answer = parse_answer(row[9], qtype)
            points = clean_str(row[11] or '')
            difficulty = DIFFICULTY_MAP.get(row[12], 'medium')
            
            q = {
                'id': f"{subject_enum}_{ch_num}_{len(chapters.get(ch_num, {}).get('questions', [])):04d}",
                'title': clean_str(row[3])[:40],
                'prompt': clean_str(row[3]),
                'type': qtype,
                'subject': subject_enum,
                'difficulty': difficulty,
                'options': options,
                'explanation': clean_str(row[10] or ''),
                'chapter': ch_num,
                'subsection': f'{ch_num}.0',
                'knowledgePoints': [points] if points else [],
            }
            
            if 'answerIndex' in answer:
                q['answerIndex'] = answer['answerIndex']
            if 'answerIndices' in answer:
                q['answerIndices'] = answer['answerIndices']
            if 'isCorrect' in answer:
                q['isCorrect'] = answer['isCorrect']
            
            if is_exam:
                if ch_name not in exam_groups:
                    exam_groups[ch_name] = []
                exam_groups[ch_name].append(q)
            else:
                if ch_num not in chapters:
                    chapters[ch_num] = {'name': ch_name, 'questions': []}
                chapters[ch_num]['questions'].append(q)
    
    # Write per-chapter JSON files
    subject_dir = os.path.join(output_dir, subject_enum)
    os.makedirs(subject_dir, exist_ok=True)
    
    total = 0
    index = {}
    
    for ch_num in sorted(chapters.keys(), key=lambda x: int(x)):
        ch = chapters[ch_num]
        ch_questions = ch['questions']
        # Re-index for this chapter
        for i, q in enumerate(ch_questions):
            q['id'] = f"{subject_enum}_{ch_num}_{i:04d}"
        
        filename = f"ch{ch_num}.json"
        filepath = os.path.join(subject_dir, filename)
        with open(filepath, 'w', encoding='utf-8') as f:
            json.dump(ch_questions, f, ensure_ascii=False)
        size_kb = os.path.getsize(filepath) / 1024
        index[ch_num] = {'name': ch['name'], 'file': filename, 'count': len(ch_questions)}
        total += len(ch_questions)
        print(f"  Chapter {ch_num}: {len(ch_questions)} questions, {size_kb:.0f}KB")
    
    # Write exam questions as one file
    if exam_groups:
        all_exam = []
        for source_name, qs in exam_groups.items():
            for i, q in enumerate(qs):
                q['id'] = f"{subject_enum}_exam_{len(all_exam):04d}"
                all_exam.append(q)
        
        filename = "exams.json"
        filepath = os.path.join(subject_dir, filename)
        with open(filepath, 'w', encoding='utf-8') as f:
            json.dump(all_exam, f, ensure_ascii=False)
        size_kb = os.path.getsize(filepath) / 1024
        index['exams'] = {'name': '历年真题与模拟练习', 'file': filename, 'count': len(all_exam)}
        total += len(all_exam)
        print(f"  Exams: {len(all_exam)} questions, {size_kb:.0f}KB")
    
    # Write index file
    index_path = os.path.join(subject_dir, 'index.json')
    with open(index_path, 'w', encoding='utf-8') as f:
        json.dump(index, f, ensure_ascii=False)
    
    print(f"  Total: {total} questions")
    return total

# Convert all three subjects
base_dir = r'D:\Android\Android\二级建造师题库'
output_dir = r'D:\Android\Android\assets\questions'

subjects = [
    ('建设工程法规及相关知识_题库.xlsx', 'law'),
    ('建设工程施工管理_题库.xlsx', 'management'),
    ('市政公用工程管理与实务_题库.xlsx', 'practice'),
]

total = 0
for filename, enum_name in subjects:
    xlsx_path = os.path.join(base_dir, filename)
    count = convert_xlsx(xlsx_path, enum_name, output_dir)
    total += count

print(f"\n=== Grand total: {total} questions ===")
# Show total asset size
import subprocess
result = subprocess.run(['dir', '/s', output_dir], capture_output=True, text=True, shell=True)
print(f"Total asset size: check assets/questions/ directory")