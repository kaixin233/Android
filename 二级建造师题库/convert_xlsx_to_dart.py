#!/usr/bin/env python3
"""Convert 二级建造师 xlsx question banks to Dart Question objects."""

import re
import json
import os
from openpyxl import load_workbook
from collections import OrderedDict

TYPE_MAP = {
    '单项选择题': 'singleChoice',
    '多项选择题': 'multipleChoice',
    '判断题': 'trueFalse',
}

DIFFICULTY_MAP = {
    '简单': 'easy',
    '中等': 'medium',
    '困难': 'hard',
}

def extract_chapter(source_str):
    """Extract chapter from source. Returns (chapter_num, display_name, is_exam_group)"""
    if not source_str:
        return '0', '其他题目', False
    
    s = str(source_str).strip()
    
    # Check if it's a chapter-based source
    m = re.match(r'第(\d+)章\s+(.+)', s)
    if m:
        return m.group(1), s, False
    
    # Check if it's an exam/mock source
    return '0', s, True

def parse_answer(answer_str, qtype):
    """Parse answer string to Dart format"""
    answer_str = str(answer_str).strip().upper()
    
    if qtype == 'trueFalse':
        if answer_str in ('对', '正确', 'A', '√', '是', 'TRUE'):
            return {'isCorrect': True}
        return {'isCorrect': False}
    
    if qtype == 'singleChoice':
        idx = ord(answer_str[0]) - ord('A') if answer_str and answer_str[0].isalpha() else 0
        return {'answerIndex': max(0, min(idx, 3))}

    if qtype == 'multipleChoice':
        indices = []
        for ch in answer_str:
            if ch.isalpha():
                idx = ord(ch) - ord('A')
                if 0 <= idx <= 4:
                    indices.append(idx)
        return {'answerIndices': indices}
    
    return {}

def escape_string(s):
    """Escape a string for Dart code"""
    if s is None:
        return ''
    s = str(s)
    s = re.sub(r'\s+', ' ', s)
    s = s.replace('\\', '\\\\')
    s = s.replace("'", "\\'")
    s = s.replace('$', '\\$')
    return s.strip()

def generate_question_dart(q, idx, subject_enum, chapter_num='0'):
    """Generate a Dart Question() constructor call from a row"""
    qtype = TYPE_MAP.get(q[1], 'singleChoice')
    prompt = escape_string(q[3])
    opt_a = escape_string(q[4] or '')
    opt_b = escape_string(q[5] or '')
    opt_c = escape_string(q[6] or '')
    opt_d = escape_string(q[7] or '')
    opt_e = escape_string(q[8] or '')
    answer = parse_answer(q[9], qtype)
    explanation = escape_string(q[10] or '')
    points = escape_string(q[11] or '')
    difficulty = DIFFICULTY_MAP.get(q[12], 'medium')
    
    # Build options list
    options = []
    if opt_a: options.append(opt_a)
    if opt_b: options.append(opt_b)
    if opt_c: options.append(opt_c)
    if opt_d: options.append(opt_d)
    if opt_e and qtype == 'multipleChoice': options.append(opt_e)
    
    # Generate ID
    qid = f"{subject_enum}_ch{chapter_num}_{idx}"
    
    # Build title from first 30 chars of prompt
    title = prompt[:30] if prompt else '题目'
    
    lines = []
    lines.append(f"    Question(")
    lines.append(f"      id: '{qid}',")
    lines.append(f"      title: '{title}',")
    lines.append(f"      prompt: '{prompt}',")
    lines.append(f"      type: QuestionType.{qtype},")
    lines.append(f"      subject: QuestionSubject.{subject_enum},")
    lines.append(f"      difficulty: QuestionDifficulty.{difficulty},")
    
    if options:
        opts_str = ', '.join([f"'{o}'" for o in options])
        lines.append(f"      options: [{opts_str}],")
    else:
        lines.append(f"      options: [],")
    
    if 'answerIndex' in answer:
        lines.append(f"      answerIndex: {answer['answerIndex']},")
    if 'answerIndices' in answer:
        indices_str = ', '.join(str(i) for i in answer['answerIndices'])
        lines.append(f"      answerIndices: [{indices_str}],")
    if 'isCorrect' in answer:
        lines.append(f"      isCorrect: {str(answer['isCorrect']).lower()},")
    
    lines.append(f"      explanation: '{explanation}',")
    lines.append(f"      chapter: '{chapter_num}',")
    lines.append(f"      subsection: '{chapter_num}.0',")
    
    if points:
        lines.append(f"      knowledgePoints: ['{points}'],")
    else:
        lines.append(f"      knowledgePoints: [],")
    
    lines.append(f"    ),")
    return '\n'.join(lines)

def convert_xlsx(xlsx_path, subject_enum, output_path):
    """Convert an xlsx file to a Dart file"""
    print(f"Reading: {xlsx_path}")
    wb = load_workbook(xlsx_path)
    ws = wb.active
    
    subject_name = ws.title
    
    # Collect all questions
    questions = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[3] and row[1] and str(row[1]).strip() in TYPE_MAP:
            questions.append(row)
    
    print(f"  Total valid questions: {len(questions)}")
    
    # Group by chapter or exam source
    chapters = OrderedDict()  # chapter_num -> {'name': ..., 'questions': [...]}
    exam_groups = OrderedDict()  # source_name -> [...]
    
    for q in questions:
        ch_num, ch_name, is_exam = extract_chapter(q[2])
        if is_exam:
            if ch_name not in exam_groups:
                exam_groups[ch_name] = []
            exam_groups[ch_name].append(q)
        else:
            if ch_num not in chapters:
                chapters[ch_num] = {'name': ch_name, 'questions': []}
            chapters[ch_num]['questions'].append(q)
    
    print(f"  Chapters: {sorted(chapters.keys(), key=lambda x: int(x))}")
    print(f"  Exam groups: {len(exam_groups)}")
    
    # Write Dart file
    with open(output_path, 'w', encoding='utf-8') as f:
        f.write(f"// Auto-generated from {os.path.basename(xlsx_path)}\n")
        f.write(f"// Subject: {subject_name}\n")
        f.write(f"// Total questions: {len(questions)}\n")
        f.write(f"// Generated by convert_xlsx_to_dart.py\n\n")
        f.write(f"import '../models/question.dart';\n\n")
        f.write(f"/// {subject_name} 题库\n")
        f.write(f"class Questions{subject_enum.capitalize()} {{\n")
        f.write(f"  Questions{subject_enum.capitalize()}._();\n\n")
        f.write(f"  static final List<Question> all = [\n")
        
        idx = 0
        
        # Write chapter-organized questions first
        for ch_num in sorted(chapters.keys(), key=lambda x: int(x)):
            ch = chapters[ch_num]
            f.write(f"\n    // ============ {ch['name']} ============\n")
            for q in ch['questions']:
                idx += 1
                dart_code = generate_question_dart(q, idx, subject_enum, ch_num)
                f.write(dart_code + '\n')
        
        # Write exam/真题 questions grouped by year
        if exam_groups:
            f.write(f"\n    // ============ 历年真题与模拟练习 ============\n")
            for source_name in exam_groups:
                qs = exam_groups[source_name]
                f.write(f"\n    // --- {source_name} ({len(qs)}题) ---\n")
                for q in qs:
                    idx += 1
                    dart_code = generate_question_dart(q, idx, subject_enum)
                    f.write(dart_code + '\n')
        
        f.write(f"  ];\n")
        f.write(f"}}\n")
    
    print(f"  Written to: {output_path}")
    print(f"  Total questions output: {idx}")
    return len(questions)

# Convert all three subjects
base_dir = r'D:\Android\Android\二级建造师题库'
output_dir = r'D:\Android\Android\lib\data'

subjects = [
    ('建设工程法规及相关知识_题库.xlsx', 'law'),
    ('建设工程施工管理_题库.xlsx', 'management'),
    ('市政公用工程管理与实务_题库.xlsx', 'practice'),
]

total = 0
for filename, enum_name in subjects:
    xlsx_path = os.path.join(base_dir, filename)
    output_path = os.path.join(output_dir, f'questions_{enum_name}.dart')
    count = convert_xlsx(xlsx_path, enum_name, output_path)
    total += count

print(f"\n=== Total questions across all subjects: {total} ===")